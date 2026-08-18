"""Governed historical-lab ingestion and multi-endpoint edit evidence.

The attached Menin workbook is a convenient human-facing *wide* table, while
modeling needs one observation per row.  This module performs that conversion
without exposing source compound names by default.  It deliberately keeps
censoring, units, assay context, provenance conflicts, and split roles visible
so a qualified value such as ``>30 uM`` is never silently treated as an exact
measurement.

The loader is read-only.  It returns in-memory tables and does not write raw or
processed private data to the repository.
"""

from __future__ import annotations

import hashlib
import hmac
import json
import math
import re
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from types import MappingProxyType
from typing import Any

import numpy as np
import pandas as pd
from menin_discovery.chemistry import standardize_smiles

from .chemistry import fragment_single_cuts, tanimoto_similarity

_ALLOWED_COHORT_ROLES = frozenset({"train", "development", "locked_external", "prospective_blind"})
_EVIDENCE_ROLES = frozenset({"train", "development"})
_VALID_RELATIONS = frozenset({"=", "~", "<", "<=", ">", ">="})
_MISSING_TOKENS = frozenset(
    {"", "na", "n/a", "nan", "none", "nt", "nd", "not tested", "not determined", "blq"}
)
_RELATION_ALIASES = {
    "≤": "<=",
    "≥": ">=",
    "=<": "<=",
    "=>": ">=",
    "≈": "~",
}


@dataclass(frozen=True)
class ParsedValue:
    """One parsed numeric result with its original concentration relation."""

    relation: str
    value: float
    raw: str

    @property
    def is_exact(self) -> bool:
        return self.relation == "="

    @property
    def is_censored(self) -> bool:
        return self.relation in {"<", "<=", ">", ">="}


@dataclass(frozen=True)
class HistoricalLabDataset:
    """Pseudonymous compound and observation tables plus an audit summary."""

    compounds: pd.DataFrame
    observations: pd.DataFrame
    issues: pd.DataFrame
    summary: Mapping[str, Any] = field(default_factory=lambda: MappingProxyType({}))


@dataclass(frozen=True)
class _EndpointColumn:
    source_column: str
    endpoint: str
    unit: str
    transform: str
    measurement_type: str
    assay_family: str
    target: str = ""
    cell_line: str = ""
    species: str = ""
    route: str = ""
    directionality: str = "context_dependent"
    unit_inferred: bool = False


_ENDPOINT_COLUMNS: tuple[_EndpointColumn, ...] = (
    _EndpointColumn(
        "Binding IC50 (nM)",
        "menin_biochemical_pIC50",
        "nM",
        "pIC50",
        "IC50",
        "biochemical_binding",
        target="Menin/MEN1",
        directionality="higher_is_more_potent",
    ),
    _EndpointColumn(
        "Slope",
        "menin_binding_hill_slope",
        "unitless",
        "identity",
        "Hill slope",
        "biochemical_binding",
        target="Menin/MEN1",
    ),
    _EndpointColumn(
        "MV4;11 IC50 (nM)",
        "mv411_cellular_pIC50",
        "nM",
        "pIC50",
        "IC50",
        "cellular_functional",
        cell_line="MV4;11",
        directionality="higher_is_more_potent",
    ),
    _EndpointColumn(
        "MOLM13 IC50 (nM)",
        "molm13_cellular_pIC50",
        "nM",
        "pIC50",
        "IC50",
        "cellular_functional",
        cell_line="MOLM13",
        directionality="higher_is_more_potent",
    ),
    _EndpointColumn(
        "HL60 IC50 (nM)",
        "hl60_cellular_pIC50",
        "nM",
        "pIC50",
        "IC50",
        "cellular_selectivity_control",
        cell_line="HL60",
        directionality="lower_pIC50_is_more_selective",
    ),
    _EndpointColumn(
        "hERG IC50 (µM)",
        "herg_pIC50",
        "uM",
        "pIC50",
        "IC50",
        "electrophysiology_functional",
        target="KCNH2/hERG",
        directionality="lower_pIC50_is_safer",
    ),
    _EndpointColumn(
        "Rat PK: T1/2 (PO) h",
        "rat_po_half_life_log10_h",
        "h",
        "log10",
        "terminal half-life",
        "in_vivo_pk",
        species="rat",
        route="PO",
    ),
    _EndpointColumn(
        "Rat PK: Tmax (PO) h",
        "rat_po_tmax_log10_h",
        "h",
        "log10",
        "Tmax",
        "in_vivo_pk",
        species="rat",
        route="PO",
    ),
    _EndpointColumn(
        "Rat PK: Cmax (PO) ng/mL",
        "rat_po_cmax_log10_ng_ml",
        "ng/mL",
        "log10",
        "Cmax",
        "in_vivo_pk",
        species="rat",
        route="PO",
        directionality="higher_is_more_exposure",
    ),
    _EndpointColumn(
        "Rat PK: AUC0-t (PO)",
        "rat_po_auc0_t_log10_ng_h_ml",
        "ng*h/mL",
        "log10",
        "AUC0-t",
        "in_vivo_pk",
        species="rat",
        route="PO",
        directionality="higher_is_more_exposure",
        unit_inferred=True,
    ),
    _EndpointColumn(
        "Rat PK: AUC0-inf (PO)",
        "rat_po_auc0_inf_log10_ng_h_ml",
        "ng*h/mL",
        "log10",
        "AUC0-inf",
        "in_vivo_pk",
        species="rat",
        route="PO",
        directionality="higher_is_more_exposure",
        unit_inferred=True,
    ),
    _EndpointColumn(
        "Rat PK: AUC0-t (IV)",
        "rat_iv_auc0_t_log10_ng_h_ml",
        "ng*h/mL",
        "log10",
        "AUC0-t",
        "in_vivo_pk",
        species="rat",
        route="IV",
        directionality="higher_is_more_exposure",
        unit_inferred=True,
    ),
    _EndpointColumn(
        "Rat PK: AUC0-inf (IV)",
        "rat_iv_auc0_inf_log10_ng_h_ml",
        "ng*h/mL",
        "log10",
        "AUC0-inf",
        "in_vivo_pk",
        species="rat",
        route="IV",
        directionality="higher_is_more_exposure",
        unit_inferred=True,
    ),
    _EndpointColumn(
        "Rat PK: CL (IV) mL/kg/min",
        "rat_iv_clearance_log10_ml_kg_min",
        "mL/kg/min",
        "log10",
        "systemic clearance",
        "in_vivo_pk",
        species="rat",
        route="IV",
        directionality="lower_is_lower_clearance",
    ),
    _EndpointColumn(
        "Rat PK: Vdss (IV) L/kg",
        "rat_iv_vdss_log10_l_kg",
        "L/kg",
        "log10",
        "Vdss",
        "in_vivo_pk",
        species="rat",
        route="IV",
    ),
    _EndpointColumn(
        "Rat PK: %F",
        "rat_bioavailability_log10_percent",
        "%",
        "log10",
        "bioavailability",
        "in_vivo_pk",
        species="rat",
        route="PO",
        directionality="higher_is_more_bioavailable",
    ),
)

_DESCRIPTOR_COLUMNS = {
    "MW": "mw",
    "TPSA": "tpsa",
    "cLogP (RDKit Crippen)": "clogp",
    "pKa1 (most basic)": "pka1_predicted",
    "pKa2 (2nd most basic)": "pka2_predicted",
}


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if bool(pd.isna(value)):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _secret_bytes(secret: str | bytes) -> bytes:
    if isinstance(secret, str):
        encoded = secret.encode("utf-8")
    elif isinstance(secret, bytes):
        encoded = secret
    else:
        raise TypeError("pseudonymization_key must be text or bytes")
    if len(encoded) < 16:
        raise ValueError("pseudonymization_key must contain at least 16 bytes")
    return encoded


def _pseudonym(value: object, *, key: bytes, namespace: str, prefix: str) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    digest = hmac.new(
        key,
        f"menin-edit-data-v1\0{namespace}\0{text}".encode(),
        hashlib.sha256,
    ).hexdigest()
    return f"{prefix}-{digest[:20].upper()}"


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def parse_qualified_value(value: object) -> ParsedValue | None:
    """Parse a numeric result while preserving ``<``, ``>``, ``~`` semantics.

    Missing/non-numeric tokens are returned as ``None``.  Composite strings
    such as PK dose pairs or hERG inhibition panels are intentionally handled
    by endpoint-specific parsers instead of guessed here.
    """

    if isinstance(value, (int, float, np.number)) and not isinstance(value, bool):
        number = float(value)
        if math.isfinite(number):
            return ParsedValue("=", number, _clean_text(value))
        return None
    raw = _clean_text(value)
    if raw.casefold() in _MISSING_TOKENS:
        return None
    normalized = raw.replace(",", "").replace("−", "-")
    for original, replacement in _RELATION_ALIASES.items():
        normalized = normalized.replace(original, replacement)
    match = re.fullmatch(
        r"\s*(<=|>=|<|>|~|=)?\s*([-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?)\s*",
        normalized,
    )
    if match is None:
        return None
    relation = match.group(1) or "="
    number = float(match.group(2))
    if relation not in _VALID_RELATIONS or not math.isfinite(number):
        return None
    return ParsedValue(relation, number, raw)


def _model_scale(
    value: float,
    relation: str,
    *,
    unit: str,
    transform: str,
) -> tuple[float, float, float]:
    """Return model threshold/value and optional lower/upper bounds."""

    if transform == "identity":
        model_value = float(value)
        lower = model_value if relation in {"=", "~", ">", ">="} else np.nan
        upper = model_value if relation in {"=", "~", "<", "<="} else np.nan
        return model_value, float(lower), float(upper)
    if value <= 0:
        return np.nan, np.nan, np.nan
    if transform == "log10":
        model_value = float(math.log10(value))
        lower = model_value if relation in {"=", "~", ">", ">="} else np.nan
        upper = model_value if relation in {"=", "~", "<", "<="} else np.nan
        return model_value, float(lower), float(upper)
    if transform != "pIC50":
        raise ValueError(f"Unsupported model transform: {transform}")
    molar_factor = {"pM": 1e-12, "nM": 1e-9, "uM": 1e-6, "mM": 1e-3, "M": 1.0}.get(unit)
    if molar_factor is None:
        return np.nan, np.nan, np.nan
    model_value = float(-math.log10(value * molar_factor))
    # Concentration relations invert on the pIC50 scale.
    lower = model_value if relation in {"=", "~", "<", "<="} else np.nan
    upper = model_value if relation in {"=", "~", ">", ">="} else np.nan
    return model_value, float(lower), float(upper)


def _hERG_percent_values(value: object) -> list[tuple[ParsedValue, float]]:
    text = _clean_text(value).replace("µ", "u").replace("μ", "u")
    if not text:
        return []
    matches = re.findall(
        r"([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*%\s*@\s*"
        r"([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*uM",
        text,
        flags=re.IGNORECASE,
    )
    return [
        (ParsedValue("=", float(percent), text), float(concentration)) for percent, concentration in matches
    ]


def _rat_dose_pairs(value: object) -> list[tuple[float, float]]:
    text = _clean_text(value)
    if not text:
        return []
    pairs: list[tuple[float, float]] = []
    for segment in text.split("|"):
        match = re.fullmatch(
            r"\s*([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*/\s*"
            r"([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*",
            segment,
        )
        if match is not None:
            pair = (float(match.group(1)), float(match.group(2)))
            if pair not in pairs:
                pairs.append(pair)
    return pairs


def _provenance_conflict_index(workbook: Any) -> dict[tuple[str, str], bool]:
    if "Provenance" not in workbook.sheetnames:
        return {}
    rows = list(workbook["Provenance"].iter_rows(values_only=True))
    if not rows:
        return {}
    header = {_clean_text(value): index for index, value in enumerate(rows[0])}
    required = {"Compound", "Parameter", "Raw value(s)"}
    if not required.issubset(header):
        return {}
    result: dict[tuple[str, str], bool] = {}
    for row in rows[1:]:
        compound = _clean_text(row[header["Compound"]])
        parameter = _clean_text(row[header["Parameter"]])
        raw_values = _clean_text(row[header["Raw value(s)"]])
        if not compound or not parameter or not raw_values:
            continue
        normalized: set[str] = set()
        for token in raw_values.split("|"):
            clean = re.sub(r"\([^)]*\)", "", token).replace(" ", "").casefold()
            if clean:
                normalized.add(clean)
        result[(compound, parameter)] = len(normalized) > 1
    return result


def _raw_observation(
    *,
    raw_compound_id: str,
    compound_id: str,
    structure_id: str,
    parsed: ParsedValue,
    spec: _EndpointColumn,
    key: bytes,
    cohort_role: str,
    provenance_conflict: bool,
    dose_context: str,
    test_concentration_um: float | None = None,
) -> dict[str, Any]:
    source_row_id = _pseudonym(
        f"{raw_compound_id}\0{spec.source_column}\0{test_concentration_um or ''}",
        key=key,
        namespace="observation_source",
        prefix="ISROW",
    )
    assay_context = {
        "endpoint": spec.endpoint,
        "target": spec.target,
        "cell_line": spec.cell_line,
        "species": spec.species,
        "route": spec.route,
        "test_concentration_um": test_concentration_um,
    }
    assay_id = _pseudonym(
        json.dumps(assay_context, sort_keys=True),
        key=key,
        namespace="assay",
        prefix="IASSAY",
    )
    model_value, model_lower, model_upper = _model_scale(
        parsed.value,
        parsed.relation,
        unit=spec.unit,
        transform=spec.transform,
    )
    return {
        "source_row_id": source_row_id,
        "compound_id": compound_id,
        "structure_id": structure_id,
        "endpoint": spec.endpoint,
        "measurement_type": spec.measurement_type,
        "assay_family": spec.assay_family,
        "assay_id": assay_id,
        "target": spec.target,
        "cell_line": spec.cell_line,
        "species": spec.species,
        "route": spec.route,
        "test_concentration_um": test_concentration_um,
        "value": float(parsed.value),
        "unit": spec.unit,
        "relation": parsed.relation,
        "is_exact": parsed.is_exact,
        "is_censored": parsed.is_censored,
        "model_value": model_value,
        "model_lower": model_lower,
        "model_upper": model_upper,
        "model_transform": spec.transform,
        "directionality": spec.directionality,
        "unit_inferred": bool(spec.unit_inferred),
        "dose_context": dose_context,
        "provenance_conflict": bool(provenance_conflict),
        "source_scope": "private",
        "split_role": cohort_role,
    }


def _values_conflict(group: pd.DataFrame, *, log10_threshold: float) -> bool:
    if len(group) < 2:
        return False
    relations = set(group["relation"].astype(str))
    if len(relations) > 1:
        return True
    values = pd.to_numeric(group["model_value"], errors="coerce").dropna().to_numpy(float)
    if len(values) < 2:
        return False
    transform = str(group["model_transform"].iloc[0])
    tolerance = 10.0 if transform == "identity" and group["unit"].iloc[0] == "%" else log10_threshold
    return bool(float(np.max(values) - np.min(values)) > tolerance)


def _aggregate_observations(
    raw: pd.DataFrame,
    *,
    key: bytes,
    conflict_log10_threshold: float,
) -> pd.DataFrame:
    columns = [
        "observation_id",
        "compound_id",
        "structure_id",
        "endpoint",
        "measurement_type",
        "assay_family",
        "assay_id",
        "target",
        "cell_line",
        "species",
        "route",
        "test_concentration_um",
        "value",
        "value_min",
        "value_max",
        "unit",
        "relation",
        "is_exact",
        "is_censored",
        "model_value",
        "model_lower",
        "model_upper",
        "model_transform",
        "directionality",
        "unit_inferred",
        "dose_context",
        "replicate_count",
        "source_record_count",
        "mixed_censoring",
        "provenance_conflict",
        "label_conflict",
        "source_scope",
        "split_role",
    ]
    if raw.empty:
        return pd.DataFrame(columns=columns)
    group_columns = [
        "compound_id",
        "structure_id",
        "endpoint",
        "measurement_type",
        "assay_family",
        "assay_id",
        "target",
        "cell_line",
        "species",
        "route",
        "test_concentration_um",
        "unit",
        "model_transform",
        "directionality",
        "unit_inferred",
        "source_scope",
        "split_role",
    ]
    records: list[dict[str, Any]] = []
    for keys, group in raw.groupby(group_columns, dropna=False, sort=True):
        metadata = dict(zip(group_columns, keys, strict=True))
        exact = group[group["relation"].eq("=")]
        if not exact.empty:
            selected = exact
            relation = "="
        else:
            selected = group
            relations = set(group["relation"].astype(str))
            if len(relations) == 1:
                relation = next(iter(relations))
            else:
                relation = "~"
        values = pd.to_numeric(selected["value"], errors="coerce").dropna()
        if values.empty:
            continue
        if relation in {">", ">="}:
            value = float(values.max())
        elif relation in {"<", "<="}:
            value = float(values.min())
        else:
            value = float(values.median())
        model_value, model_lower, model_upper = _model_scale(
            value,
            relation,
            unit=str(metadata["unit"]),
            transform=str(metadata["model_transform"]),
        )
        identity = "\0".join(
            str(metadata[name])
            for name in (
                "structure_id",
                "endpoint",
                "assay_id",
                "test_concentration_um",
                "split_role",
            )
        )
        provenance_conflict = bool(group["provenance_conflict"].astype(bool).any())
        label_conflict = provenance_conflict or _values_conflict(
            group, log10_threshold=conflict_log10_threshold
        )
        dose_values = sorted(
            {value for value in group["dose_context"].astype(str) if value and value != "nan"}
        )
        records.append(
            {
                "observation_id": _pseudonym(identity, key=key, namespace="observation", prefix="IOBS"),
                **metadata,
                "value": value,
                "value_min": float(pd.to_numeric(group["value"], errors="coerce").min()),
                "value_max": float(pd.to_numeric(group["value"], errors="coerce").max()),
                "relation": relation,
                "is_exact": relation == "=",
                "is_censored": relation in {"<", "<=", ">", ">="},
                "model_value": model_value,
                "model_lower": model_lower,
                "model_upper": model_upper,
                "dose_context": " | ".join(dose_values),
                "replicate_count": int(len(group)),
                "source_record_count": int(group["source_row_id"].nunique()),
                "mixed_censoring": int(group["relation"].nunique()) > 1,
                "provenance_conflict": provenance_conflict,
                "label_conflict": label_conflict,
            }
        )
    return pd.DataFrame(records, columns=columns).sort_values(
        ["structure_id", "endpoint", "test_concentration_um"],
        kind="stable",
        ignore_index=True,
        na_position="last",
    )


def load_historical_lab_workbook(
    path: str | Path,
    *,
    pseudonymization_key: str | bytes,
    cohort_role: str = "development",
    include_raw_identifiers: bool = False,
    conflict_log10_threshold: float = 0.5,
) -> HistoricalLabDataset:
    """Load the historical wide workbook into governed long-form tables.

    ``pseudonymization_key`` is required and must be supplied at runtime.  Raw
    source compound names are omitted unless ``include_raw_identifiers=True``
    is explicitly requested.  No files are written by this function.
    """

    role = _clean_text(cohort_role).casefold()
    if role not in _ALLOWED_COHORT_ROLES:
        raise ValueError(f"Unsupported cohort_role: {cohort_role!r}")
    if conflict_log10_threshold < 0 or not math.isfinite(conflict_log10_threshold):
        raise ValueError("conflict_log10_threshold must be finite and non-negative")
    key = _secret_bytes(pseudonymization_key)
    source_path = Path(path).expanduser().resolve()
    if not source_path.is_file():
        raise FileNotFoundError(source_path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional environment.
        raise ImportError("Install the Menin-Edit 'lab' extra to read .xlsx workbooks") from exc

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        if "SMILES" not in workbook.sheetnames:
            raise ValueError("Historical workbook must contain a 'SMILES' sheet")
        values = list(workbook["SMILES"].iter_rows(values_only=True))
        if not values:
            raise ValueError("Historical workbook 'SMILES' sheet is empty")
        headers = [_clean_text(value) for value in values[0]]
        if len(headers) != len(set(headers)):
            raise ValueError("Historical workbook contains duplicate column headers")
        required = {"Compound", "Kekule Canonical SMILES"}
        missing = sorted(required - set(headers))
        if missing:
            raise ValueError(f"Historical workbook is missing required columns: {missing}")
        source = pd.DataFrame(values[1:], columns=headers).dropna(how="all").reset_index(drop=True)
        provenance_conflicts = _provenance_conflict_index(workbook)
    finally:
        workbook.close()

    issues: list[dict[str, str]] = []
    structure_rows: list[dict[str, Any]] = []
    raw_observations: list[dict[str, Any]] = []
    available_specs = [spec for spec in _ENDPOINT_COLUMNS if spec.source_column in source.columns]

    for source_position, row in source.iterrows():
        raw_compound_id = _clean_text(row.get("Compound"))
        source_row_id = _pseudonym(
            raw_compound_id or f"row-{source_position + 2}",
            key=key,
            namespace="source_compound",
            prefix="ISRCMP",
        )
        raw_smiles = _clean_text(row.get("Kekule Canonical SMILES"))
        standardized = standardize_smiles(
            raw_smiles,
            strip_salts=True,
            canonicalize_tautomer=False,
            require_rdkit=True,
        )
        if not standardized.structure_valid or not standardized.structure_id:
            issues.append(
                {
                    "row_id": source_row_id,
                    "severity": "error",
                    "code": "invalid_structure",
                    "field": "Kekule Canonical SMILES",
                    "message": "Structure could not be standardized with the registered policy",
                }
            )
            continue
        compound_id = _pseudonym(
            standardized.structure_id,
            key=key,
            namespace="compound_structure",
            prefix="ICMP",
        )
        structure_record: dict[str, Any] = {
            "compound_id": compound_id,
            "source_row_id": source_row_id,
            "structure_id": standardized.structure_id,
            "smiles": standardized.standardized_smiles,
            "standard_inchi_key": standardized.standard_inchi_key,
            "structure_standardization_version": standardized.structure_standardization_version,
            "cohort_role": role,
        }
        if include_raw_identifiers:
            structure_record["source_compound_id"] = raw_compound_id
        for source_name, target_name in _DESCRIPTOR_COLUMNS.items():
            structure_record[target_name] = pd.to_numeric(row.get(source_name), errors="coerce")
        structure_rows.append(structure_record)

        dose_context = _clean_text(row.get("Rat PK: Dose (IV/PO) mg/kg"))
        for spec in available_specs:
            raw_value = row.get(spec.source_column)
            if _clean_text(raw_value).casefold() in _MISSING_TOKENS:
                continue
            parsed = parse_qualified_value(raw_value)
            if parsed is None:
                issues.append(
                    {
                        "row_id": source_row_id,
                        "severity": "warning",
                        "code": "unparsed_value",
                        "field": spec.source_column,
                        "message": "Non-empty result could not be parsed as a qualified numeric value",
                    }
                )
                continue
            if parsed.value <= 0 and spec.transform in {"pIC50", "log10"}:
                issues.append(
                    {
                        "row_id": source_row_id,
                        "severity": "warning",
                        "code": "nonpositive_log_value",
                        "field": spec.source_column,
                        "message": "A positive value is required for the registered model transform",
                    }
                )
                continue
            raw_observations.append(
                _raw_observation(
                    raw_compound_id=raw_compound_id,
                    compound_id=compound_id,
                    structure_id=standardized.structure_id,
                    parsed=parsed,
                    spec=spec,
                    key=key,
                    cohort_role=role,
                    provenance_conflict=provenance_conflicts.get(
                        (raw_compound_id, spec.source_column), False
                    ),
                    dose_context=dose_context if spec.species == "rat" else "",
                )
            )

        if "hERG % inhibition" in source.columns:
            inhibition_raw = row.get("hERG % inhibition")
            parsed_inhibition = _hERG_percent_values(inhibition_raw)
            if _clean_text(inhibition_raw) and not parsed_inhibition:
                issues.append(
                    {
                        "row_id": source_row_id,
                        "severity": "warning",
                        "code": "unparsed_herg_percent_panel",
                        "field": "hERG % inhibition",
                        "message": "hERG percent-inhibition text did not match the registered grammar",
                    }
                )
            percent_spec = _EndpointColumn(
                "hERG % inhibition",
                "herg_percent_inhibition",
                "%",
                "identity",
                "percent inhibition",
                "electrophysiology_functional",
                target="KCNH2/hERG",
                directionality="lower_is_safer",
            )
            for parsed, concentration in parsed_inhibition:
                raw_observations.append(
                    _raw_observation(
                        raw_compound_id=raw_compound_id,
                        compound_id=compound_id,
                        structure_id=standardized.structure_id,
                        parsed=parsed,
                        spec=percent_spec,
                        key=key,
                        cohort_role=role,
                        provenance_conflict=provenance_conflicts.get(
                            (raw_compound_id, "hERG % inhibition"), False
                        ),
                        dose_context="",
                        test_concentration_um=concentration,
                    )
                )

        # Dose is modeled as explicit IV and PO context observations only when
        # the ``IV/PO`` pair grammar is unambiguous.
        dose_pairs = _rat_dose_pairs(dose_context)
        for pair_index, (iv_dose, po_dose) in enumerate(dose_pairs, start=1):
            for route_name, dose in (("IV", iv_dose), ("PO", po_dose)):
                dose_spec = _EndpointColumn(
                    "Rat PK: Dose (IV/PO) mg/kg",
                    f"rat_{route_name.casefold()}_dose_mg_kg",
                    "mg/kg",
                    "identity",
                    "dose",
                    "in_vivo_pk_context",
                    species="rat",
                    route=route_name,
                )
                raw_observations.append(
                    _raw_observation(
                        raw_compound_id=f"{raw_compound_id}\0dose-pair-{pair_index}",
                        compound_id=compound_id,
                        structure_id=standardized.structure_id,
                        parsed=ParsedValue("=", dose, dose_context),
                        spec=dose_spec,
                        key=key,
                        cohort_role=role,
                        provenance_conflict=len(dose_pairs) > 1,
                        dose_context=dose_context,
                    )
                )

    structures = pd.DataFrame(structure_rows)
    raw_frame = pd.DataFrame(raw_observations)
    observations = _aggregate_observations(
        raw_frame,
        key=key,
        conflict_log10_threshold=conflict_log10_threshold,
    )

    compound_columns = [
        "compound_id",
        "structure_id",
        "smiles",
        "standard_inchi_key",
        "structure_standardization_version",
        "cohort_role",
        "source_record_count",
        "duplicate_structure",
        "descriptor_conflict",
        "label_conflict",
        "conflict_flag",
        *_DESCRIPTOR_COLUMNS.values(),
    ]
    if include_raw_identifiers:
        compound_columns.append("source_compound_ids")
    compound_records: list[dict[str, Any]] = []
    if not structures.empty:
        for structure_id, group in structures.groupby("structure_id", sort=True):
            record: dict[str, Any] = {
                "compound_id": str(group["compound_id"].iloc[0]),
                "structure_id": str(structure_id),
                "smiles": str(group["smiles"].iloc[0]),
                "standard_inchi_key": str(group["standard_inchi_key"].iloc[0]),
                "structure_standardization_version": str(group["structure_standardization_version"].iloc[0]),
                "cohort_role": role,
                "source_record_count": int(group["source_row_id"].nunique()),
                "duplicate_structure": int(group["source_row_id"].nunique()) > 1,
            }
            descriptor_conflict = False
            for descriptor in _DESCRIPTOR_COLUMNS.values():
                values_numeric = pd.to_numeric(group[descriptor], errors="coerce").dropna()
                record[descriptor] = float(values_numeric.median()) if not values_numeric.empty else np.nan
                if values_numeric.nunique() > 1:
                    descriptor_conflict = True
            label_conflict = bool(
                observations.loc[observations["structure_id"].eq(structure_id), "label_conflict"]
                .astype(bool)
                .any()
            )
            record["descriptor_conflict"] = descriptor_conflict
            record["label_conflict"] = label_conflict
            record["conflict_flag"] = descriptor_conflict or label_conflict
            if include_raw_identifiers:
                record["source_compound_ids"] = ";".join(sorted(set(group["source_compound_id"].astype(str))))
            compound_records.append(record)
    compounds = pd.DataFrame(compound_records, columns=compound_columns)

    issue_columns = ["row_id", "severity", "code", "field", "message"]
    issue_frame = pd.DataFrame(issues, columns=issue_columns).sort_values(
        ["severity", "code", "row_id"], kind="stable", ignore_index=True
    )
    workbook_hash = _file_sha256(source_path)
    summary = MappingProxyType(
        {
            "schema_version": "1.0",
            "workbook_sha256": workbook_hash,
            "cohort_role": role,
            "raw_identifiers_included": bool(include_raw_identifiers),
            "source_rows": int(len(source)),
            "standardized_compounds": int(len(compounds)),
            "duplicate_structure_groups": int(compounds["duplicate_structure"].sum())
            if not compounds.empty
            else 0,
            "conflicted_structure_groups": int(compounds["conflict_flag"].sum())
            if not compounds.empty
            else 0,
            "observations": int(len(observations)),
            "endpoint_counts": {
                str(endpoint): int(count)
                for endpoint, count in observations["endpoint"].value_counts().sort_index().items()
            },
            "censored_observations": int(observations["is_censored"].sum()) if not observations.empty else 0,
            "provenance_conflicts": int(observations["provenance_conflict"].sum())
            if not observations.empty
            else 0,
            "issue_counts": {
                str(code): int(count)
                for code, count in issue_frame["code"].value_counts().sort_index().items()
            },
        }
    )
    return HistoricalLabDataset(
        compounds=compounds,
        observations=observations,
        issues=issue_frame,
        summary=summary,
    )


def _delta_bounds(left: pd.Series, right: pd.Series) -> tuple[float, float, float]:
    exact = bool(left["is_exact"]) and bool(right["is_exact"])
    point = float(right["model_value"] - left["model_value"]) if exact else np.nan
    left_lower = pd.to_numeric(pd.Series([left["model_lower"]]), errors="coerce").iloc[0]
    left_upper = pd.to_numeric(pd.Series([left["model_upper"]]), errors="coerce").iloc[0]
    right_lower = pd.to_numeric(pd.Series([right["model_lower"]]), errors="coerce").iloc[0]
    right_upper = pd.to_numeric(pd.Series([right["model_upper"]]), errors="coerce").iloc[0]
    lower = float(right_lower - left_upper) if pd.notna(right_lower) and pd.notna(left_upper) else np.nan
    upper = float(right_upper - left_lower) if pd.notna(right_upper) and pd.notna(left_lower) else np.nan
    return point, lower, upper


def build_multi_endpoint_mmp_evidence(
    compounds: pd.DataFrame,
    observations: pd.DataFrame,
    *,
    allowed_split_roles: Sequence[str] = ("train", "development"),
    min_core_heavy_atoms: int = 10,
    max_variable_heavy_atoms: int = 12,
    exact_only: bool = True,
    include_conflicts: bool = False,
    bidirectional: bool = True,
) -> pd.DataFrame:
    """Build governed, vector-valued matched-pair evidence.

    Only ``train`` and ``development`` observations can enter the evidence
    table.  Requests attempting to admit locked/prospective roles are rejected,
    which prevents a caller from accidentally contaminating edit discovery.
    Censored observations can optionally be retained as delta bounds, but no
    point delta is fabricated for them.
    """

    requested_roles = {_clean_text(role).casefold() for role in allowed_split_roles}
    if not requested_roles or not requested_roles.issubset(_EVIDENCE_ROLES):
        raise ValueError("MMP evidence may use only 'train' and 'development' rows")
    compound_required = {"structure_id", "smiles"}
    observation_required = {
        "structure_id",
        "endpoint",
        "model_value",
        "model_lower",
        "model_upper",
        "is_exact",
        "label_conflict",
        "split_role",
    }
    missing_compounds = sorted(compound_required - set(compounds.columns))
    missing_observations = sorted(observation_required - set(observations.columns))
    if missing_compounds:
        raise KeyError(f"Compound table is missing required columns: {missing_compounds}")
    if missing_observations:
        raise KeyError(f"Observation table is missing required columns: {missing_observations}")
    columns = [
        "pair_id",
        "structure_id_a",
        "structure_id_b",
        "core_smiles",
        "source_fragment",
        "target_fragment",
        "transformation",
        "endpoint",
        "value_a",
        "value_b",
        "delta",
        "delta_lower",
        "delta_upper",
        "relation_a",
        "relation_b",
        "context_similarity",
        "source_scope",
        "split_role",
        "evidence_grade",
    ]
    if compounds.empty or observations.empty:
        return pd.DataFrame(columns=columns)

    obs = observations.copy()
    obs["split_role"] = obs["split_role"].fillna("").astype(str).str.casefold()
    obs = obs[obs["split_role"].isin(requested_roles)]
    if not include_conflicts:
        obs = obs[~obs["label_conflict"].fillna(False).astype(bool)]
    if exact_only:
        obs = obs[obs["is_exact"].fillna(False).astype(bool)]
    if obs.empty:
        return pd.DataFrame(columns=columns)

    # Reduce repeated assay rows only inside the already-governed role subset.
    # The workbook loader normally emits one row here; this also supports public
    # long tables adapted to the same contract.
    endpoint_rows: dict[str, dict[str, pd.Series]] = {}
    for (structure_id, endpoint), group in obs.groupby(["structure_id", "endpoint"], sort=True):
        values = pd.to_numeric(group["model_value"], errors="coerce")
        finite = group.loc[np.isfinite(values)]
        if finite.empty and exact_only:
            continue
        endpoint_rows.setdefault(str(structure_id), {})[str(endpoint)] = (
            finite.iloc[0] if not finite.empty else group.iloc[0]
        )

    eligible_ids = set(endpoint_rows)
    molecule_table = (
        compounds.loc[compounds["structure_id"].astype(str).isin(eligible_ids), ["structure_id", "smiles"]]
        .drop_duplicates("structure_id")
        .sort_values("structure_id", kind="stable")
        .reset_index(drop=True)
    )
    if len(molecule_table) < 2:
        return pd.DataFrame(columns=columns)

    fragment_index: dict[str, list[tuple[str, str, int, int]]] = {}
    smiles_by_id = dict(
        zip(molecule_table["structure_id"].astype(str), molecule_table["smiles"].astype(str), strict=True)
    )
    for row in molecule_table.itertuples(index=False):
        for context in fragment_single_cuts(
            str(row.smiles),
            min_core_heavy_atoms=min_core_heavy_atoms,
            max_variable_heavy_atoms=max_variable_heavy_atoms,
        ):
            fragment_index.setdefault(context.core_smiles, []).append(
                (
                    str(row.structure_id),
                    context.variable_smiles,
                    int(context.core_heavy_atoms),
                    int(context.variable_heavy_atoms),
                )
            )

    # One molecule pair can share several cuts; use the largest retained core
    # for a single, deterministic explanation.
    pair_contexts: dict[tuple[str, str], tuple[int, str, str, str]] = {}
    for core, items_raw in sorted(fragment_index.items()):
        items = sorted(set(items_raw))
        for left_index, left in enumerate(items):
            for right in items[left_index + 1 :]:
                left_id, left_fragment, left_core_heavy, _ = left
                right_id, right_fragment, right_core_heavy, _ = right
                if left_id == right_id or left_fragment == right_fragment:
                    continue
                if left_id > right_id:
                    left_id, right_id = right_id, left_id
                    left_fragment, right_fragment = right_fragment, left_fragment
                candidate = (
                    min(left_core_heavy, right_core_heavy),
                    core,
                    left_fragment,
                    right_fragment,
                )
                key_pair = (left_id, right_id)
                if key_pair not in pair_contexts or candidate[0] > pair_contexts[key_pair][0]:
                    pair_contexts[key_pair] = candidate

    records: list[dict[str, Any]] = []
    for (left_id, right_id), (_core_heavy, core, left_fragment, right_fragment) in sorted(
        pair_contexts.items()
    ):
        shared_endpoints = sorted(set(endpoint_rows.get(left_id, {})) & set(endpoint_rows.get(right_id, {})))
        if not shared_endpoints:
            continue
        similarity = tanimoto_similarity(smiles_by_id[left_id], smiles_by_id[right_id])
        for endpoint in shared_endpoints:
            left = endpoint_rows[left_id][endpoint]
            right = endpoint_rows[right_id][endpoint]
            delta, lower, upper = _delta_bounds(left, right)
            if exact_only and not math.isfinite(delta):
                continue
            left_role = _clean_text(left.get("split_role", "development")).casefold()
            right_role = _clean_text(right.get("split_role", "development")).casefold()
            split_role = "train" if left_role == right_role == "train" else "development"
            left_assay = _clean_text(left.get("assay_id", ""))
            right_assay = _clean_text(right.get("assay_id", ""))
            evidence_grade = "same_assay" if left_assay and left_assay == right_assay else "same_series"
            source_scope_values = {
                _clean_text(left.get("source_scope", "private")) or "private",
                _clean_text(right.get("source_scope", "private")) or "private",
            }
            source_scope = next(iter(source_scope_values)) if len(source_scope_values) == 1 else "mixed"

            directions: Iterable[tuple[str, str, str, str, pd.Series, pd.Series, float, float, float]]
            forward = (
                left_id,
                right_id,
                left_fragment,
                right_fragment,
                left,
                right,
                delta,
                lower,
                upper,
            )
            if bidirectional:
                reverse_delta = -delta if math.isfinite(delta) else np.nan
                reverse_lower = -upper if math.isfinite(upper) else np.nan
                reverse_upper = -lower if math.isfinite(lower) else np.nan
                directions = (
                    forward,
                    (
                        right_id,
                        left_id,
                        right_fragment,
                        left_fragment,
                        right,
                        left,
                        reverse_delta,
                        reverse_lower,
                        reverse_upper,
                    ),
                )
            else:
                directions = (forward,)
            for (
                structure_a,
                structure_b,
                source_fragment,
                target_fragment,
                row_a,
                row_b,
                signed_delta,
                signed_lower,
                signed_upper,
            ) in directions:
                pair_digest = hashlib.sha256(
                    f"{structure_a}\0{structure_b}\0{core}\0{endpoint}".encode()
                ).hexdigest()[:20]
                records.append(
                    {
                        "pair_id": f"PAIR-{pair_digest.upper()}",
                        "structure_id_a": structure_a,
                        "structure_id_b": structure_b,
                        "core_smiles": core,
                        "source_fragment": source_fragment,
                        "target_fragment": target_fragment,
                        "transformation": f"{source_fragment}>>{target_fragment}",
                        "endpoint": endpoint,
                        "value_a": float(row_a["model_value"]),
                        "value_b": float(row_b["model_value"]),
                        "delta": signed_delta,
                        "delta_lower": signed_lower,
                        "delta_upper": signed_upper,
                        "relation_a": _clean_text(row_a.get("relation", "=")) or "=",
                        "relation_b": _clean_text(row_b.get("relation", "=")) or "=",
                        "context_similarity": float(similarity),
                        "source_scope": source_scope,
                        "split_role": split_role,
                        "evidence_grade": evidence_grade,
                    }
                )
    if not records:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(records, columns=columns).sort_values(
        ["endpoint", "source_fragment", "target_fragment", "structure_id_a", "structure_id_b"],
        kind="stable",
        ignore_index=True,
    )


__all__ = [
    "HistoricalLabDataset",
    "ParsedValue",
    "build_multi_endpoint_mmp_evidence",
    "load_historical_lab_workbook",
    "parse_qualified_value",
]
